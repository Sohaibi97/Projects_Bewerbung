from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def adjust_columns(file_path):
    # I load the workbookfile and select the first worksheet
    wb = load_workbook(file_path)
    ws = wb.worksheets[0]

    # i loop through each column and adjust the width based on the max length in that column
    for column in ws.columns:
        max_length = 0 # first the max width is 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: # str() is important here because not all columns have cells
                    max_length = len(cell.value) # and now i change my max value to the true max value
            except:
                pass # the exceptions happens because of cells without a value
        adjusted_width = (max_length + 2) # i want my cells to be a bit bigger than their data therefore +2
        # get_column_letter verwandelt die Spaltennummer (1, 2, ..) zu A, B, denn ws.column_dimensions kann nur
        # mit Buchstaben umgehen.Mit ws.column_dimenstions kann ich die Breite mit .width einstellen
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width 

    # i save the file and Basile can look at it without problems
    wb.save(file_path)
